"""
Database Services for Yemen ID Documents.

Provides separate SQLite databases for:
- Yemen National ID Cards
- Yemen Passports

Features:
- CRUD operations
- CSV/Excel export
- Name parsing (first/middle/last)
"""
import csv
import sqlite3
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from abc import ABC, abstractmethod

# Base directories
BASE_DIR = Path(__file__).parent.parent
DATABASE_DIR = BASE_DIR / "data" / "databases"
EXPORTS_DIR = BASE_DIR / "data" / "exports"

# Ensure directories exist
DATABASE_DIR.mkdir(parents=True, exist_ok=True)
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def split_name(full_name: str, is_arabic: bool = False) -> Dict[str, str]:
    """
    Split a full name into first, middle, and last name components.
    
    Arabic names typically follow: First + Father's Name + Grandfather's Name + Family Name
    English names typically follow: First + Middle + Last
    
    Args:
        full_name: The complete name string
        is_arabic: Whether the name is in Arabic
        
    Returns:
        Dictionary with first_name, middle_name, last_name
    """
    if not full_name or not full_name.strip():
        return {
            "first_name": "",
            "middle_name": "",
            "last_name": ""
        }
    
    # Clean and split the name
    parts = full_name.strip().split()
    
    if len(parts) == 1:
        return {
            "first_name": parts[0],
            "middle_name": "",
            "last_name": ""
        }
    elif len(parts) == 2:
        return {
            "first_name": parts[0],
            "middle_name": "",
            "last_name": parts[1]
        }
    elif len(parts) == 3:
        return {
            "first_name": parts[0],
            "middle_name": parts[1],
            "last_name": parts[2]
        }
    else:
        # For longer names: first is first, last is last, everything else is middle
        return {
            "first_name": parts[0],
            "middle_name": " ".join(parts[1:-1]),
            "last_name": parts[-1]
        }


class BaseDatabase(ABC):
    """Abstract base class for document databases."""
    
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self._create_table()
    
    def _get_connection(self) -> sqlite3.Connection:
        """Get database connection with row factory."""
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        return conn
    
    @abstractmethod
    def _create_table(self):
        """Create the database table if it doesn't exist."""
        pass
    
    @abstractmethod
    def get_table_name(self) -> str:
        """Return the table name."""
        pass
    
    @abstractmethod
    def get_columns(self) -> List[str]:
        """Return list of column names for export."""
        pass
    
    def get_all(self) -> List[Dict[str, Any]]:
        """Get all records from the database."""
        conn = self._get_connection()
        try:
            cursor = conn.execute(f"SELECT * FROM {self.get_table_name()} ORDER BY created_at DESC")
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()
    
    def get_by_id(self, record_id: int) -> Optional[Dict[str, Any]]:
        """Get a record by its primary key ID."""
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                f"SELECT * FROM {self.get_table_name()} WHERE id = ?",
                (record_id,)
            )
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()
    
    def delete(self, record_id: int) -> bool:
        """Delete a record by its primary key ID."""
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                f"DELETE FROM {self.get_table_name()} WHERE id = ?",
                (record_id,)
            )
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()
    
    def count(self) -> int:
        """Get total number of records."""
        conn = self._get_connection()
        try:
            cursor = conn.execute(f"SELECT COUNT(*) FROM {self.get_table_name()}")
            return cursor.fetchone()[0]
        finally:
            conn.close()
    
    def export_csv(self, filename: Optional[str] = None) -> Path:
        """
        Export all records to a CSV file.
        
        Args:
            filename: Optional filename, defaults to table_name_YYYYMMDD_HHMMSS.csv
            
        Returns:
            Path to the exported CSV file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.get_table_name()}_{timestamp}.csv"
        
        export_path = EXPORTS_DIR / filename
        records = self.get_all()
        columns = self.get_columns()
        
        with open(export_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(records)
        
        return export_path
    
    def export_excel(self, filename: Optional[str] = None) -> Path:
        """
        Export all records to an Excel file.
        
        Args:
            filename: Optional filename, defaults to table_name_YYYYMMDD_HHMMSS.xlsx
            
        Returns:
            Path to the exported Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.get_table_name()}_{timestamp}.xlsx"
        
        export_path = EXPORTS_DIR / filename
        records = self.get_all()
        columns = self.get_columns()
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.get_table_name()
        
        # Write headers
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = cell.font.copy(bold=True)
        
        # Write data
        for row_idx, record in enumerate(records, 2):
            for col_idx, column in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(column, ""))
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(columns, 1):
            max_length = len(column)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        wb.save(export_path)
        return export_path


class YemenIDCardDB(BaseDatabase):
    """Database for Yemen National ID Cards."""
    
    def __init__(self):
        super().__init__(DATABASE_DIR / "yemen_id_cards.db")
    
    def get_table_name(self) -> str:
        return "id_cards"
    
    def get_columns(self) -> List[str]:
        return [
            "id", "national_id",
            "first_name_arabic", "middle_name_arabic", "last_name_arabic",
            "first_name_english", "middle_name_english", "last_name_english",
            "date_of_birth", "place_of_birth", "gender", "blood_group",
            "issuance_date", "expiry_date",
            "front_image_blob", "back_image_blob", "selfie_image_blob",
            "created_at"
        ]
    
    def _create_table(self):
        """Create the id_cards table if it doesn't exist."""
        conn = self._get_connection()
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS id_cards (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    national_id TEXT UNIQUE NOT NULL,
                    first_name_arabic TEXT,
                    middle_name_arabic TEXT,
                    last_name_arabic TEXT,
                    first_name_english TEXT,
                    middle_name_english TEXT,
                    last_name_english TEXT,
                    date_of_birth TEXT,
                    place_of_birth TEXT,
                    gender TEXT,
                    blood_group TEXT,
                    issuance_date TEXT,
                    expiry_date TEXT,
                    front_image_blob BLOB,
                    back_image_blob BLOB,
                    selfie_image_blob BLOB,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.commit()
        finally:
            conn.close()
    
    def insert(self, data: Dict[str, Any]) -> int:
        """
        Insert a new ID card record.
        
        Args:
            data: Dictionary with ID card data. Can include:
                - national_id (required)
                - name_arabic (will be split into first/middle/last)
                - name_english (will be split into first/middle/last)
                - Or individual first_name_*, middle_name_*, last_name_* fields
                - date_of_birth, place_of_birth, gender, etc.
                - front_image_blob, back_image_blob, selfie_image_blob (bytes)
                
        Returns:
            The inserted record ID
        """
        # Parse names if provided as full names
        if "name_arabic" in data:
            name_parts = split_name(data["name_arabic"], is_arabic=True)
            data["first_name_arabic"] = name_parts["first_name"]
            data["middle_name_arabic"] = name_parts["middle_name"]
            data["last_name_arabic"] = name_parts["last_name"]
        
        if "name_english" in data:
            name_parts = split_name(data["name_english"], is_arabic=False)
            data["first_name_english"] = name_parts["first_name"]
            data["middle_name_english"] = name_parts["middle_name"]
            data["last_name_english"] = name_parts["last_name"]
        
        conn = self._get_connection()
        try:
            # Use local time for created_at instead of SQLite's UTC CURRENT_TIMESTAMP
            local_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor = conn.execute("""
                INSERT INTO id_cards (
                    national_id, 
                    first_name_arabic, middle_name_arabic, last_name_arabic,
                    first_name_english, middle_name_english, last_name_english,
                    date_of_birth, place_of_birth, gender, blood_group,
                    issuance_date, expiry_date,
                    front_image_blob, back_image_blob, selfie_image_blob,
                    created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get("national_id"),
                data.get("first_name_arabic"), data.get("middle_name_arabic"), data.get("last_name_arabic"),
                data.get("first_name_english"), data.get("middle_name_english"), data.get("last_name_english"),
                data.get("date_of_birth"), data.get("place_of_birth"), data.get("gender"), data.get("blood_group"),
                data.get("issuance_date"), data.get("expiry_date"),
                data.get("front_image_blob"), data.get("back_image_blob"), data.get("selfie_image_blob"),
                local_timestamp
            ))
            conn.commit()
            return cursor.lastrowid
        finally:
            conn.close()
    
    def get_by_national_id(self, national_id: str) -> Optional[Dict[str, Any]]:
        """Get a record by national ID number."""
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                "SELECT * FROM id_cards WHERE national_id = ?",
                (national_id,)
            )
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()
    
    def update(self, national_id: str, data: Dict[str, Any]) -> bool:
        """Update an existing ID card record by national ID."""
        # Parse names if provided as full names
        if "name_arabic" in data:
            name_parts = split_name(data["name_arabic"], is_arabic=True)
            data["first_name_arabic"] = name_parts["first_name"]
            data["middle_name_arabic"] = name_parts["middle_name"]
            data["last_name_arabic"] = name_parts["last_name"]
        
        if "name_english" in data:
            name_parts = split_name(data["name_english"], is_arabic=False)
            data["first_name_english"] = name_parts["first_name"]
            data["middle_name_english"] = name_parts["middle_name"]
            data["last_name_english"] = name_parts["last_name"]
        
        # Build dynamic UPDATE query
        update_fields = []
        values = []
        
        allowed_fields = [
            "first_name_arabic", "middle_name_arabic", "last_name_arabic",
            "first_name_english", "middle_name_english", "last_name_english",
            "date_of_birth", "place_of_birth", "gender", "blood_group",
            "issuance_date", "expiry_date",
            "front_image_blob", "back_image_blob", "selfie_image_blob"
        ]
        
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = ?")
                values.append(data[field])
        
        if not update_fields:
            return False
        
        values.append(national_id)
        
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                f"UPDATE id_cards SET {', '.join(update_fields)} WHERE national_id = ?",
                values
            )
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()


class YemenPassportDB(BaseDatabase):
    """Database for Yemen Passports."""
    
    def __init__(self):
        super().__init__(DATABASE_DIR / "yemen_passports.db")
    
    def get_table_name(self) -> str:
        return "passports"
    
    def get_columns(self) -> List[str]:
        return [
            "id", "passport_number",
            "surname_arabic", "given_names_arabic",
            "surname_english", "given_names_english",
            "profession",
            "date_of_birth", "place_of_birth", "gender", "blood_group",
            "passport_type", "issuance_date", "expiry_date", "issuing_authority",
            "mrz_line_1", "mrz_line_2",
            "passport_image_blob", "selfie_image_blob",
            "created_at"
        ]
    
    def _create_table(self):
        """Create the passports table if it doesn't exist."""
        conn = self._get_connection()
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS passports (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    passport_number TEXT UNIQUE NOT NULL,
                    surname_arabic TEXT,
                    given_names_arabic TEXT,
                    surname_english TEXT,
                    given_names_english TEXT,
                    profession TEXT,
                    date_of_birth TEXT,
                    place_of_birth TEXT,
                    gender TEXT,
                    blood_group TEXT,
                    passport_type TEXT DEFAULT 'Ordinary',
                    issuance_date TEXT,
                    expiry_date TEXT,
                    issuing_authority TEXT,
                    mrz_line_1 TEXT,
                    mrz_line_2 TEXT,
                    passport_image_blob BLOB,
                    selfie_image_blob BLOB,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.commit()
        finally:
            conn.close()
    
    def insert(self, data: Dict[str, Any]) -> int:
        """
        Insert a new passport record.
        
        Args:
            data: Dictionary with passport data. Can include:
                - passport_number (required)
                - surname_arabic, given_names_arabic
                - surname_english, given_names_english
                - profession
                - date_of_birth, place_of_birth, gender, blood_group, etc.
                
        Returns:
            The inserted record ID
        """
        conn = self._get_connection()
        try:
            # Use local time for created_at instead of SQLite's UTC CURRENT_TIMESTAMP
            local_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor = conn.execute("""
                INSERT INTO passports (
                    passport_number,
                    surname_arabic, given_names_arabic,
                    surname_english, given_names_english,
                    profession,
                    date_of_birth, place_of_birth, gender, blood_group,
                    passport_type, issuance_date, expiry_date, issuing_authority,
                    mrz_line_1, mrz_line_2,
                    passport_image_blob, selfie_image_blob,
                    created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get("passport_number"),
                data.get("surname_arabic"), data.get("given_names_arabic"),
                data.get("surname_english"), data.get("given_names_english"),
                data.get("profession"),
                data.get("date_of_birth"), data.get("place_of_birth"), data.get("gender"), data.get("blood_group"),
                data.get("passport_type", "Ordinary"),
                data.get("issuance_date"), data.get("expiry_date"),
                data.get("issuing_authority"),
                data.get("mrz_line_1"), data.get("mrz_line_2"),
                data.get("passport_image_blob"), data.get("selfie_image_blob"),
                local_timestamp
            ))
            conn.commit()
            return cursor.lastrowid
        finally:
            conn.close()
    
    def get_by_passport_number(self, passport_number: str) -> Optional[Dict[str, Any]]:
        """Get a record by passport number."""
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                "SELECT * FROM passports WHERE passport_number = ?",
                (passport_number,)
            )
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()
    
    def update(self, passport_number: str, data: Dict[str, Any]) -> bool:
        """Update an existing passport record by passport number."""
        # Build dynamic UPDATE query
        update_fields = []
        values = []
        
        allowed_fields = [
            "surname_arabic", "given_names_arabic",
            "surname_english", "given_names_english",
            "profession",
            "date_of_birth", "place_of_birth", "gender", "blood_group",
            "passport_type", "issuance_date", "expiry_date", "issuing_authority",
            "mrz_line_1", "mrz_line_2",
            "passport_image_blob", "selfie_image_blob"
        ]
        
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = ?")
                values.append(data[field])
        
        if not update_fields:
            return False
        
        values.append(passport_number)
        
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                f"UPDATE passports SET {', '.join(update_fields)} WHERE passport_number = ?",
                values
            )
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()


class VerificationDB(BaseDatabase):
    """Database for storing verification results."""
    
    def __init__(self):
        super().__init__(DATABASE_DIR / "verification_results.db")
    
    def get_table_name(self) -> str:
        return "verifications"
    
    def get_columns(self) -> List[str]:
        return [
            "id", "document_type", "document_id",
            # ID Card name fields (used when document_type = 'id_card')
            "first_name_arabic", "middle_name_arabic", "last_name_arabic",
            "first_name_english", "middle_name_english", "last_name_english",
            # Passport name fields (used when document_type = 'passport')
            "surname_arabic", "given_names_arabic",
            "surname_english", "given_names_english",
            "profession",
            # Common fields
            "date_of_birth", "place_of_birth", "gender", "blood_group",
            # Verification results
            "verification_status", "similarity_score",
            # Images
            "selfie_image_blob", "document_image_blob",
            # Timestamps
            "verified_at", "created_at"
        ]
    
    def _create_table(self):
        """Create the verifications table if it doesn't exist."""
        conn = self._get_connection()
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS verifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    document_type TEXT NOT NULL,
                    document_id TEXT NOT NULL,
                    first_name_arabic TEXT,
                    middle_name_arabic TEXT,
                    last_name_arabic TEXT,
                    first_name_english TEXT,
                    middle_name_english TEXT,
                    last_name_english TEXT,
                    surname_arabic TEXT,
                    given_names_arabic TEXT,
                    surname_english TEXT,
                    given_names_english TEXT,
                    profession TEXT,
                    date_of_birth TEXT,
                    place_of_birth TEXT,
                    gender TEXT,
                    blood_group TEXT,
                    verification_status TEXT DEFAULT 'pending',
                    similarity_score REAL,
                    selfie_image_blob BLOB,
                    document_image_blob BLOB,
                    verified_at TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.commit()
        finally:
            conn.close()
    
    def insert(self, data: Dict[str, Any]) -> int:
        """
        Insert a new verification record.
        
        Args:
            data: Dictionary with verification data. Required fields:
                - document_type ('id_card' or 'passport')
                - document_id (national_id or passport_number)
                - verification_status ('verified', 'failed', 'pending')
                - similarity_score (0.0 - 1.0)
                
        Returns:
            The inserted record ID
        """
        conn = self._get_connection()
        try:
            local_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            verified_at = local_timestamp if data.get("verification_status") == "verified" else None
            
            cursor = conn.execute("""
                INSERT INTO verifications (
                    document_type, document_id,
                    first_name_arabic, middle_name_arabic, last_name_arabic,
                    first_name_english, middle_name_english, last_name_english,
                    surname_arabic, given_names_arabic,
                    surname_english, given_names_english,
                    profession,
                    date_of_birth, place_of_birth, gender, blood_group,
                    verification_status, similarity_score,
                    selfie_image_blob, document_image_blob,
                    verified_at, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get("document_type"),
                data.get("document_id"),
                data.get("first_name_arabic"), data.get("middle_name_arabic"), data.get("last_name_arabic"),
                data.get("first_name_english"), data.get("middle_name_english"), data.get("last_name_english"),
                data.get("surname_arabic"), data.get("given_names_arabic"),
                data.get("surname_english"), data.get("given_names_english"),
                data.get("profession"),
                data.get("date_of_birth"), data.get("place_of_birth"), data.get("gender"), data.get("blood_group"),
                data.get("verification_status", "pending"), data.get("similarity_score"),
                data.get("selfie_image_blob"), data.get("document_image_blob"),
                verified_at, local_timestamp
            ))
            conn.commit()
            return cursor.lastrowid
        finally:
            conn.close()
    
    def get_by_document_id(self, document_id: str, document_type: str = None) -> List[Dict[str, Any]]:
        """Get all verification records for a document ID."""
        conn = self._get_connection()
        try:
            if document_type:
                cursor = conn.execute(
                    "SELECT * FROM verifications WHERE document_id = ? AND document_type = ? ORDER BY created_at DESC",
                    (document_id, document_type)
                )
            else:
                cursor = conn.execute(
                    "SELECT * FROM verifications WHERE document_id = ? ORDER BY created_at DESC",
                    (document_id,)
                )
            return [dict(row) for row in cursor.fetchall()]
        finally:
            conn.close()
    
    def get_verified_records(self) -> List[Dict[str, Any]]:
        """Get all successful verification records."""
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                "SELECT * FROM verifications WHERE verification_status = 'verified' ORDER BY verified_at DESC"
            )
            return [dict(row) for row in cursor.fetchall()]
        finally:
            conn.close()
    
    def update(self, record_id: int, data: Dict[str, Any]) -> bool:
        """Update an existing verification record by ID."""
        update_fields = []
        values = []
        
        allowed_fields = [
            "verification_status", "similarity_score",
            "selfie_image_blob", "document_image_blob", "verified_at"
        ]
        
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = ?")
                values.append(data[field])
        
        if not update_fields:
            return False
        
        values.append(record_id)
        
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                f"UPDATE verifications SET {', '.join(update_fields)} WHERE id = ?",
                values
            )
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()


# Singleton instances
_id_card_db: Optional[YemenIDCardDB] = None
_passport_db: Optional[YemenPassportDB] = None
_verification_db: Optional[VerificationDB] = None


def get_id_card_db() -> YemenIDCardDB:
    """Get the Yemen ID Card database instance."""
    global _id_card_db
    if _id_card_db is None:
        _id_card_db = YemenIDCardDB()
    return _id_card_db


def get_passport_db() -> YemenPassportDB:
    """Get the Yemen Passport database instance."""
    global _passport_db
    if _passport_db is None:
        _passport_db = YemenPassportDB()
    return _passport_db


def get_verification_db() -> VerificationDB:
    """Get the Verification database instance."""
    global _verification_db
    if _verification_db is None:
        _verification_db = VerificationDB()
    return _verification_db
